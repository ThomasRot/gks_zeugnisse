# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# Struktur: (Fach, [ (Kompetenz, [Subkompetenzen...]) ... ])
# Kompetenz "" = keine seitliche Gruppierung (z. B. Kunst, Musik)
PLACEHOLDER = "(Kompetenzen hier ergänzen)"

faecher = [
    ("Deutsch", [
        ("Sprechen und Zuhören", [
            "Du gibst Informationen korrekt wieder.",
            "Du beziehst dich auf die Beiträge anderer.",
            "Du kannst grammatikalisch richtige Sätze bilden.",
        ]),
        ("Lesen und Texte rezipieren", [
            "Du kannst altersgemäße Texte in angemessener Zeit lesen und verstehen.",
            "Du kannst Fragen zum Gelesenen mündlich und schriftlich beantworten.",
            "Du markierst wichtige Aussagen eines Textes passend.",
            "Du kannst den Inhalt des Gelesenen in Stichworten notieren.",
            "Du kannst wichtige Inhalte des Gelesenen erfassen und mithilfe von Textstellen belegen.",
            "Du äußerst zu altersgemäßen Texten eigene Gedanken und Einschätzungen.",
            "Du kannst geübte Texte flüssig und mit Betonung vorlesen.",
        ]),
        ("Schreiben", [
            "Du schreibst lesbar und zügig.",
            "Du wendest Rechtschreibstrategien auch bei unbekannten Wörtern richtig an.",
            "Du schreibst geübte Wörter aus dem Grundwortschatz richtig.",
            "Du schreibst Nomen und Satzanfänge groß.",
            "Du hältst Wort- und Satzgrenzen ein.",
            "Du gliederst deine Texte übersichtlich.",
            "Du schreibst fehlerfrei ab.",
            "Du kannst das Wörterbuch als Rechtschreibhilfe nutzen.",
            "Du kannst vollständige, grammatikalisch richtige Sätze schreiben.",
            "Du schreibst Geschichten, Berichte und Personenbeschreibungen nach besprochenen Kriterien.",
            "Du entwickelst Schreibideen für Texte.",
            "Du gestaltest eigene Texte mit Hilfe sprachlicher Mittel (z. B. wechselnde Satzanfänge, passende Zeitformen, wörtliche Rede).",
            "Du reflektierst und überarbeitest deine Texte.",
            "Du präsentierst deine Texte vor anderen.",
        ]),
        ("Sprache untersuchen", [
            "Du ordnest Wörter den Wortarten Nomen, Adjektive, Verben, Artikel, Präpositionen und Pronomen passend zu.",
            "Du unterscheidest Aussage-, Frage- und Ausrufesatz und verwendest die passenden Satzschlusszeichen.",
            "Du erkennst die wörtliche Rede und setzt die passenden Satzzeichen.",
            "Du benennst und bildest Verbformen in verschiedenen Zeiten (Präsens, Präteritum, Perfekt) richtig.",
            "Du kannst Wörter nach dem Alphabet sortieren und nachschlagen.",
        ]),
    ]),
    ("Mathematik", [
        ("Zahl und Operation (Zahlenraum bis 1.000)", [
            "Du kannst dich im Zahlenraum bis 1000 orientieren.",
            "Du kannst Additionsaufgaben (+) mit und ohne Übergänge lösen.",
            "Du kannst Additionsaufgaben halbschriftlich und schriftlich lösen.",
            "Du kannst Subtraktionsaufgaben (-) mit und ohne Übergänge lösen.",
            "Du kannst Subtraktionsaufgaben mit Übertrag halbschriftlich und schriftlich lösen.",
            "Du löst die Multiplikationsaufgaben (*) des kleinen Einmaleins automatisiert.",
            "Du löst die Divisionsaufgaben (:) des kleinen Einmaleins automatisiert.",
            "Du kannst Divisionsaufgaben (:) mit Rest lösen.",
            "Du wählst geeignete Rechenstrategien aus und wendest sie an.",
            "Du nutzt geeignete Verfahren, um deine Ergebnisse zu überprüfen.",
            "Du kannst Sachaufgaben verstehen und eine passende Fragestellung entwickeln.",
            "Du kannst zu Sachaufgaben Lösungswege finden und eine passende Antwort formulieren.",
        ]),
        ("Raum und Form", [
            "Du kannst geometrischen Grundformen benennen und ihre besonderen Merkmale mit Fachbegriffen beschreiben.",
            "Du kannst geometrische Formen mit und ohne Lineal zeichnen.",
            "Du vergrößerst oder verkleinerst Figuren nach Vorgaben.",
            "Du erkennst komplexere geometrische Muster und kannst diese fortsetzen.",
            "Du kannst geometrische Körper benennen, mit Fachbegriffen beschreiben sie vergleichen und ihnen die entsprechenden Körpernetze zuordnen.",
            "Du überprüfst Figuren auf Achsensymmetrie und kannst zur Spiegelachse symmetrisch ergänzen.",
            "Du kannst Lagebeziehungen und Ansichten beschreiben.",
            "Du kannst Lagepläne lesen und zeichnen sowie Wege auf Lageplänen finden, beschreiben und zeichnen.",
        ]),
        ("Größen und Messen", [
            "Du kannst Uhrzeiten sekundengenau ablesen, einstellen und notieren.",
            "Du kannst Zeitspannen bestimmen.",
            "Du kannst Repräsentanten zu den Längeneinheiten mm, cm, m und km benennen und Längen abmessen.",
            "Du kannst Längenangaben in unterschiedlichen Schreibweisen darstellen.",
            "Du kannst Repräsentanten zu den Gewichtseinheiten g, kg und t benennen und Objekte wiegen.",
            "Du kannst Gewichtsangaben in unterschiedlichen Schreibweisen darstellen.",
            "Du kannst Geldbeträge mit Scheinen und Münzen darstellen, miteinander vergleichen und ordnen.",
            "Du kannst mit Geldbeträgen rechnen.",
        ]),
        ("Daten und Zufall", [
            "Du kannst Daten aus einer Tabelle oder einem Diagramm ablesen bzw. selbst erheben und in eine andere Darstellungsform übertragen.",
            "Du kannst kombinatorische Aufgaben durch systematisches Vorgehen lösen und deine Lösung strategisch darstellen.",
            "Du kannst Wahrscheinlichkeiten von einfachen Zufallsversuchen einschätzen, die Ergebnisse beschreiben und miteinander vergleichen.",
        ]),
        ("Problemlösen, Modellieren, Argumentieren, Darstellen/Kommunizieren", [
            "Du stellst deine Ideen verständlich dar und kannst deinen Lösungsweg erklären sowie mit anderen vergleichen.",
            "Du kannst Begründungen und Argumente formulieren.",
        ]),
    ]),
    ("Sachunterricht", [
        ("Erkenntnisgewinnung", [
            "Du kannst zielführend Fragen stellen, um Wissen zu erwerben und Sachthemen zu verstehen.",
            "Du kannst zu Sachthemen recherchieren, indem du verschiedene Quellen (Sachtexte, Interviews, Bild-, Text- und Sachquellen) auswertest.",
            "Du kannst einen Versuch (zum Pflanzenwachstum, zum Sprudelgas, zur magnetischen Wirkung von Strom) planen, aufbauen, durchführen und auswerten.",
            "Du kannst Daten zu (untersuchten Bereich einfügen) erheben, darstellen und auswerten.",
            "Du kannst (Stadtpläne, Baupläne, mehrschrittige Anleitungen, PC-Programme, Apps) lesen und nutzen.",
            "Du wendest Problemlösestrategien (im Klassenrat, bei Interessenskonflikten, bei technischen Fragestellungen) an.",
            "Du kannst (Modelle, technische Konstruktionen) (planen, bauen, erklären) und dadurch Erkenntnisse über (Fahrzeuge, Türme, Brücken, Pflanzen, Körperteile) gewinnen.",
        ]),
        ("Kommunikation", [
            "Du bringst dich in Unterrichtsgesprächen aktiv ein und zeigst dadurch dein Interesse an Sachthemen.",
            "Du kannst Beobachtungen und Vermutungen von/ zu (Pflanzen, Tieren, Bildern, Gegenständen) äußern.",
            "Du verwendest treffende Begriffe und sachlich angemessene Darstellungen (Symbole, Zeichnungen) zu den Sachthemen (Unterrichtsthemen einfügen).",
            "Du nutzt geeignete Präsentations- und Darstellungsformen nach besprochenen Kriterien.",
        ]),
        ("Bewertung", [
            "Du kannst mit anderen darüber diskutieren, ob Dargestelltes real oder fiktiv ist.",
            "Du kannst (Ereignisse, Gegenstände) in der Zeit einordnen.",
            "Du benennst Maßnahmen zur Erhaltung der eigenen Gesundheit (Zähne/ gesunde Ernährung/ mein Körper).",
            "Du kannst beschreiben, dass Umwelt und Natur im Sinne der Nachhaltigkeit geschützt werden müssen und gehst respektvoll mit Lebewesen um.",
            "Du kennst die Grundsätze des demokratischen Systems und wendest sie (im Klassenrat, bei Entscheidungsfindungen, politischen Diskussionen) an.",
        ]),
    ]),
    ("Englisch", [("", [PLACEHOLDER])]),
    ("Sport", [("", [PLACEHOLDER])]),
    ("Religion", [("", [PLACEHOLDER])]),
    ("Kunst", [("", [
        "Du beteiligst dich an Unterrichtsgesprächen und zeigst so Interesse an künstlerischen Themen.",
        "Du nutzt Buntstifte, Wasserfarben, Acrylfarben und selbst ausgewählte Bastelmaterialien für deine Werke.",
        "Du beachtest gestalterische Kriterien bei der Umsetzung eines Themas.",
        "Du hast eigene Ideen zur Gestaltung.",
        "Du setzt deine Ideen selbstständig um.",
        "Du stellst deine Werke in der vorgesehenen Zeit sinnvoll fertig.",
        "Du kannst Tipps annehmen und sinnvoll für die eigene Arbeit nutzen.",
        "Du kannst deine eigenen Kunstwerke und die anderer Künstler beschreiben.",
        "Du begründest deine Meinung zu Kunstwerken.",
        "Du kennst Fachbegriffe und nutzt sie passend.",
        "Du kannst ausdrücken, ob du dich an die Gestaltungsvorgaben gehalten hast.",
    ])]),
    ("Musik", [("", [
        "Du beteiligst dich interessiert an den musikalischen Aktivitäten.",
        "Du beteiligst dich aktiv an Unterrichtsgesprächen.",
        "Du findest passende Bewegungen und entwickelst Choreografien zu unterschiedlicher Musik.",
        "Du kannst Rhythmen umsetzen.",
        "Du singst unsere Lieder mit.",
        "Du tanzt vorgegebene Tänze nach.",
        "Du kennst alle Orff-Instrumente und spielst sie rhythmisch und technisch korrekt.",
        "Du kennst Orchester-Instrumente.",
        "Du kennst musikalische Gestaltungsmittel wie hoch-tief, laut-leise, schnell-langsam, traurig-fröhlich und kannst sie umsetzen.",
        "Du kannst Notenwerte benennen und umsetzen.",
        "Du nutzt musikalische Fachbegriffe.",
    ])]),
]


def build_rows():
    rows = [
        ["Name", ""],
        ["Schuljahr", "2025/2026"],
        ["Kommentar", ""],
        [],
        ["Fach", "Kompetenz", "Subkompetenz", "Bewertung", "Fachkommentar"],
    ]
    for fach, komps in faecher:
        first_fach_row = True
        for komp, subs in komps:
            first_komp_row = True
            for sub in subs:
                rows.append([
                    fach if first_fach_row else "",
                    komp if first_komp_row else "",
                    sub,
                    "",   # Bewertung leer -> Lehrkraft fuellt
                    "",   # Fachkommentar
                ])
                first_fach_row = False
                first_komp_row = False
    return rows


wb = Workbook()
ws = wb.active
ws.title = "Zeugnis"
for r in build_rows():
    ws.append(r)

# Formatierung (Tabellenkopf jetzt Zeile 5, Metadaten A1–A3)
for c in ws[5]:
    c.font = Font(bold=True)
for addr in ("A1", "A2", "A3"):
    ws[addr].font = Font(bold=True)
widths = {"A": 16, "B": 26, "C": 70, "D": 11, "E": 26}
for col, w in widths.items():
    ws.column_dimensions[col].width = w
for r in ws.iter_rows():
    for cell in r:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

# Blatt Bewertungsskala
hs = wb.create_sheet("Bewertungsskala")
for row in [
    ["Bewertung", "Bedeutung", "Symbol"],
    [1, "sicher", "voller Kreis ●"],
    [2, "überwiegend", "3/4-Kreis ◕"],
    [3, "teilweise", "1/2-Kreis ◑"],
    [4, "mit Unterstützung", "1/4-Kreis ◔"],
    [5, "Kompetenz wird noch erworben", "→"],
    ["*", "siehe Kommentar", "*"],
    [],
    ["Hinweise", "", ""],
    ["", "Leere Zellen bei Fach/Kompetenz = wie in der Zeile darüber.", ""],
    ["", "Bewertung leer lassen, wenn (noch) nicht beurteilt.", ""],
    ["", "Datei in Numbers über Ablage → Exportieren → Excel wieder als .xlsx speichern.", ""],
]:
    hs.append(row)
hs.column_dimensions["A"].width = 12
hs.column_dimensions["B"].width = 52
hs.column_dimensions["C"].width = 18
for c in hs[1]:
    c.font = Font(bold=True)
hs["A9"].font = Font(bold=True)

out = "/Users/A200139250/Programming/gks_zeugnisse/zeugnis-vorlage.xlsx"
wb.save(out)

# kleine Zusammenfassung
total_subs = sum(len(subs) for _, komps in faecher for _, subs in komps)
print("saved", out)
print("Faecher:", len(faecher), "| Subkompetenzen:", total_subs, "| Datenzeilen:", len(build_rows()) - 6)
